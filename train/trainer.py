from __future__ import annotations

import math
import os
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, TextIO
from zoneinfo import ZoneInfo

import openpyxl
import torch
import torch.nn.functional as F
import wandb
import yaml
from accelerate import Accelerator
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from accelerate.utils import set_seed
from torch.utils.data import DataLoader
from tqdm import tqdm
from transformers import get_scheduler

from train.config import TrainConfig
from train.data import RerankerDataset, load_eval_datasets, make_collate_fn
from train.modeling import extract_yes_no_logits, load_model_and_tokenizer

BEIJING_TZ = ZoneInfo("Asia/Shanghai")


@dataclass
class LossBreakdown:
    total: torch.Tensor
    rank: torch.Tensor
    listwise: torch.Tensor
    pointwise: torch.Tensor

    def as_float_dict(self) -> dict[str, float]:
        return {
            "loss_total": self.total.detach().item(),
            "loss_rank": self.rank.detach().item(),
            "loss_listwise": self.listwise.detach().item(),
            "loss_pointwise": self.pointwise.detach().item(),
        }


@dataclass
class TrainingState:
    best_mrr: float = float("-inf")
    global_samples_seen: int = 0
    epoch_samples_seen: int = 0
    optimizer_steps: int = 0
    next_eval_at: int = 0


class MetricTracker:
    def __init__(self) -> None:
        self.reset()

    def reset(self) -> None:
        self._totals = {
            "loss_total": 0.0,
            "loss_rank": 0.0,
            "loss_listwise": 0.0,
            "loss_pointwise": 0.0,
        }
        self.count = 0

    def update(self, losses: LossBreakdown) -> None:
        for key, value in losses.as_float_dict().items():
            self._totals[key] += value
        self.count += 1

    def averages(self) -> dict[str, float]:
        if self.count == 0:
            return {key: 0.0 for key in self._totals}
        return {key: value / self.count for key, value in self._totals.items()}

    def distributed_averages(self, accelerator: Accelerator) -> dict[str, float]:
        if self.count == 0:
            return {key: 0.0 for key in self._totals}

        stats = torch.tensor(
            [
                self._totals["loss_total"],
                self._totals["loss_rank"],
                self._totals["loss_listwise"],
                self._totals["loss_pointwise"],
                float(self.count),
            ],
            dtype=torch.float32,
            device=accelerator.device,
        )
        gathered = accelerator.gather(stats).view(-1, stats.numel()).sum(dim=0)
        total_count = max(gathered[-1].item(), 1.0)
        return {
            "loss_total": gathered[0].item() / total_count,
            "loss_rank": gathered[1].item() / total_count,
            "loss_listwise": gathered[2].item() / total_count,
            "loss_pointwise": gathered[3].item() / total_count,
        }


class TrainLogger:
    def __init__(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        self._handle: TextIO = open(path, "a", encoding="utf-8")  # noqa: SIM115

    def _timestamp(self) -> str:
        return datetime.now(BEIJING_TZ).strftime("%Y-%m-%d %H:%M:%S UTC+08:00")

    def info(self, message: str) -> None:
        self._handle.write(f"[{self._timestamp()}] {message}\n")
        self._handle.flush()

    def log_config(self, cfg: TrainConfig) -> None:
        self.info("=" * 60)
        self.info("Resolved training configuration")
        self.info("=" * 60)
        for key, value in cfg.to_flat_dict().items():
            self.info(f"  {key}: {value}")
        self.info("=" * 60)

    def log_step(
        self,
        *,
        step: int,
        epoch: int,
        samples_seen: int,
        learning_rate: float,
        losses: dict[str, float],
        grad_norm: float | None,
    ) -> None:
        parts = [
            f"step={step}",
            f"epoch={epoch}",
            f"samples={samples_seen}",
            f"lr={learning_rate:.2e}",
            f"weighted-loss-backward={losses['loss_total']:.6f}",
            f"raw_rank={losses['loss_rank']:.6f}",
            f"raw_list={losses['loss_listwise']:.6f}",
            f"raw_point={losses['loss_pointwise']:.6f}",
        ]
        if grad_norm is not None:
            parts.append(f"grad_norm={grad_norm:.4f}")
        self.info("[TRAIN] " + " | ".join(parts))

    def log_eval(
        self,
        *,
        step: int,
        samples_seen: int,
        mrr: float,
        best_mrr: float,
        is_best: bool,
        stage: str,
    ) -> None:
        suffix = " ** NEW BEST **" if is_best else ""
        self.info(
            f"[{stage.upper()}] step={step} | samples={samples_seen} | "
            f"mrr={mrr:.6f} | best_mrr={best_mrr:.6f}{suffix}"
        )

    def close(self) -> None:
        self._handle.close()


def compute_rank_loss(
    student_z: torch.Tensor,
    teacher_scores: torch.Tensor,
    hard_neg_scale: float,
    num_positives: int,
) -> torch.Tensor:
    """Rank loss: for each positive, cross-entropy against all negatives.

    When ``num_positives == 1`` this is mathematically equivalent to the
    original single-positive implementation.
    """
    pos_logits = student_z[:, :num_positives]
    neg_logits = student_z[:, num_positives:]
    pos_teacher = teacher_scores[:, :num_positives]
    neg_teacher = teacher_scores[:, num_positives:]

    batch_size = student_z.size(0)
    target = torch.zeros(batch_size, dtype=torch.long, device=student_z.device)

    total_loss = torch.tensor(0.0, device=student_z.device)
    for i in range(num_positives):
        single_pos_teacher = pos_teacher[:, i : i + 1]
        margins = (single_pos_teacher - neg_teacher).clamp(min=0)
        difficulty = 1.0 / (1.0 + margins * hard_neg_scale)
        neg_weights = F.softmax(difficulty, dim=-1)

        single_pos_logit = pos_logits[:, i : i + 1]
        weighted_neg = neg_logits + torch.log(neg_weights + 1e-8)
        combined = torch.cat([single_pos_logit, weighted_neg], dim=-1)
        total_loss = total_loss + F.cross_entropy(combined, target)

    return total_loss / num_positives


def compute_list_loss(
    student_z: torch.Tensor,
    teacher_scores: torch.Tensor,
    temperature: float,
) -> torch.Tensor:
    clamped_scores = teacher_scores.clamp(1e-6, 1 - 1e-6)
    teacher_logits = torch.log(clamped_scores / (1.0 - clamped_scores))
    teacher_probs = F.softmax(teacher_logits / temperature, dim=-1)
    student_log_probs = F.log_softmax(student_z / temperature, dim=-1)
    return F.kl_div(student_log_probs, teacher_probs, reduction="batchmean") * (
        temperature**2
    )


def compute_point_loss(
    student_z: torch.Tensor,
    teacher_scores: torch.Tensor,
) -> torch.Tensor:
    return F.mse_loss(torch.sigmoid(student_z), teacher_scores)


def compute_losses(
    student_z: torch.Tensor,
    teacher_scores: torch.Tensor,
    cfg: TrainConfig,
    num_positives: int,
) -> LossBreakdown:
    loss_rank = compute_rank_loss(
        student_z, teacher_scores, cfg.loss.hard_neg_scale, num_positives
    )
    loss_list = compute_list_loss(student_z, teacher_scores, cfg.loss.temperature)
    loss_point = compute_point_loss(student_z, teacher_scores)
    loss_total = (
        cfg.loss.alpha_rank * loss_rank
        + cfg.loss.beta_list * loss_list
        + cfg.loss.gamma_point * loss_point
    )
    return LossBreakdown(
        total=loss_total,
        rank=loss_rank,
        listwise=loss_list,
        pointwise=loss_point,
    )


@torch.no_grad()
def evaluate(
    model: Any,
    dev_loader: DataLoader,
    accelerator: Accelerator,
    micro_batch_size: int = 2,
) -> float:
    model.eval()

    # Disable gradient checkpointing during eval for faster forward pass
    unwrapped = accelerator.unwrap_model(model)
    gc_enabled = getattr(unwrapped, "is_gradient_checkpointing", False)
    if gc_enabled:
        unwrapped.gradient_checkpointing_disable()

    reciprocal_ranks: list[float] = []

    for batch in tqdm(
        dev_loader,
        desc="Eval",
        leave=False,
        disable=not accelerator.is_main_process,
    ):
        input_ids = batch["input_ids"]
        attention_mask = batch["attention_mask"]

        logit_parts: list[torch.Tensor] = []
        for chunk_ids, chunk_mask in zip(
            input_ids.split(micro_batch_size, dim=0),
            attention_mask.split(micro_batch_size, dim=0),
        ):
            logit_parts.append(extract_yes_no_logits(model, chunk_ids, chunk_mask))
        scores = torch.cat(logit_parts, dim=0).float()

        num_positives: int = batch["num_positives"]
        best_pos_score = scores[:num_positives].max().item()
        rank = 1 + (scores[num_positives:] >= best_pos_score).sum().item()
        reciprocal_ranks.append(1.0 / rank)

    if gc_enabled:
        unwrapped.gradient_checkpointing_enable()

    rr_tensor = torch.tensor(
        reciprocal_ranks,
        dtype=torch.float32,
        device=accelerator.device,
    )
    gathered = accelerator.gather_for_metrics(rr_tensor)
    model.train()
    return gathered.mean().item() if gathered.numel() > 0 else 0.0


def save_model_bundle(
    model: Any,
    tokenizer: Any,
    cfg: TrainConfig,
    save_dir: Path,
    accelerator: Accelerator,
) -> None:
    save_dir.mkdir(parents=True, exist_ok=True)
    unwrapped = accelerator.unwrap_model(model)
    unwrapped.save_pretrained(str(save_dir))
    tokenizer.save_pretrained(str(save_dir))
    with open(save_dir / "train_config.resolved.yaml", "w", encoding="utf-8") as handle:
        yaml.safe_dump(cfg.to_dict(), handle, sort_keys=False, allow_unicode=False)
    accelerator.print(f"Saved checkpoint to {save_dir}")


def save_eval_results_xlsx(
    output_path: Path,
    column_name: str,
    per_file_mrr: dict[str, float],
) -> None:
    """Append one evaluation column to the xlsx results table.

    Rows = dataset names, columns = evaluation checkpoints (e.g. samples-1000).
    """
    table = _load_eval_xlsx(output_path)
    table[column_name] = dict(per_file_mrr)

    wb = Workbook()
    ws = wb.active
    ws.title = "Eval MRR"

    col_names = list(table)
    dataset_names = sorted({ds for scores in table.values() for ds in scores})

    ws.cell(row=1, column=1, value="Dataset")
    for ci, col in enumerate(col_names, start=2):
        ws.cell(row=1, column=ci, value=col)

    for ri, ds in enumerate(dataset_names, start=2):
        ws.cell(row=ri, column=1, value=ds)
        for ci, col in enumerate(col_names, start=2):
            score = table[col].get(ds)
            if score is not None:
                ws.cell(row=ri, column=ci, value=round(score, 6))

    avg_row = len(dataset_names) + 2
    ws.cell(row=avg_row, column=1, value="AVERAGE")
    for ci, col in enumerate(col_names, start=2):
        vals = [table[col][ds] for ds in dataset_names if ds in table[col]]
        if vals:
            ws.cell(row=avg_row, column=ci, value=round(sum(vals) / len(vals), 6))

    _format_eval_sheet(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _load_eval_xlsx(path: Path) -> dict[str, dict[str, float]]:
    """Load existing evaluation xlsx into column_name -> {dataset: score}."""
    if not path.exists():
        return {}

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [
        str(ws.cell(row=1, column=c).value)
        for c in range(2, ws.max_column + 1)
        if ws.cell(row=1, column=c).value
    ]
    table: dict[str, dict[str, float]] = {h: {} for h in headers}
    for row in range(2, ws.max_row + 1):
        ds = ws.cell(row=row, column=1).value
        if not ds or str(ds) == "AVERAGE":
            continue
        for offset, header in enumerate(headers, start=2):
            val = ws.cell(row=row, column=offset).value
            if isinstance(val, (int, float)):
                table[header][str(ds)] = float(val)
    return table


def _format_eval_sheet(ws: Any) -> None:
    """Apply formatting to the evaluation results sheet."""
    header_fill = PatternFill(
        start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"
    )
    bold = Font(bold=True)

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in range(2, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        if label == "AVERAGE":
            for c in range(1, ws.max_column + 1):
                ws.cell(row=row, column=c).font = bold
            continue
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")

    ws.column_dimensions["A"].width = 40
    for c in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    for row in range(2, ws.max_row + 1):
        for c in range(2, ws.max_column + 1):
            cell = ws.cell(row=row, column=c)
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"
                cell.alignment = Alignment(horizontal="center")


class RerankerTrainer:
    def __init__(
        self, cfg: TrainConfig, *, config_path: str | Path | None = None
    ) -> None:
        self.cfg = cfg
        self.config_path = Path(config_path) if config_path else None
        self.accelerator = Accelerator(
            gradient_accumulation_steps=cfg.training.grad_accum_steps
        )
        set_seed(cfg.training.seed)

        self.output_dir = Path(cfg.output.dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        self.logger: TrainLogger | None = None
        self.wandb_run: Any | None = None
        self._init_tracking()

        self.accelerator.print(f"Loading model from {cfg.model.path} ...")
        self.model, self.tokenizer = load_model_and_tokenizer(cfg)
        self.accelerator.print("Model loaded.")

        self.train_dataset = RerankerDataset(
            cfg.data.train_file,
            max_samples=cfg.data.train_samples,
            seed=cfg.training.seed,
        )
        if not self.train_dataset:
            raise ValueError("Training dataset is empty.")

        self.dev_datasets = load_eval_datasets(
            cfg.data.dev_dir,
            max_samples=cfg.data.eval_samples,
            max_files=cfg.data.eval_files,
            seed=cfg.training.seed,
            num_neg=cfg.evaluation.num_neg,
        )

        self.accelerator.print(f"Training samples: {len(self.train_dataset)}")
        total_dev = sum(len(ds) for ds in self.dev_datasets.values())
        self.accelerator.print(
            f"Dev datasets: {len(self.dev_datasets)} files, {total_dev} total samples"
        )
        if self.logger:
            self.logger.info(f"Training samples loaded: {len(self.train_dataset)}")
            for name, ds in self.dev_datasets.items():
                self.logger.info(f"Dev dataset '{name}': {len(ds)} samples")

        collate_fn = make_collate_fn(self.tokenizer, cfg.model.max_seq_length)
        pin_memory = cfg.data.pin_memory and torch.cuda.is_available()
        self.train_loader = DataLoader(
            self.train_dataset,
            batch_size=1,
            shuffle=True,
            collate_fn=collate_fn,
            num_workers=cfg.data.num_workers,
            pin_memory=pin_memory,
            drop_last=False,
        )
        self.dev_loaders: dict[str, DataLoader] = {}
        for name, ds in self.dev_datasets.items():
            self.dev_loaders[name] = DataLoader(
                ds,
                batch_size=1,
                shuffle=False,
                collate_fn=collate_fn,
                num_workers=cfg.data.num_workers,
                pin_memory=pin_memory,
            )

        self.optimizer = torch.optim.AdamW(
            self.model.parameters(),
            lr=cfg.training.learning_rate,
            weight_decay=cfg.training.weight_decay,
        )

        queries_per_optimizer_update = (
            cfg.training.grad_accum_steps * self.accelerator.num_processes
        )
        self.optimizer_updates_per_epoch = max(
            1,
            math.ceil(len(self.train_dataset) / queries_per_optimizer_update),
        )
        # Accelerate already adjusts scheduler stepping for distributed training,
        # so the schedule should only be reduced by grad accumulation here.
        self.scheduler_steps_per_epoch = max(
            1,
            math.ceil(len(self.train_dataset) / cfg.training.grad_accum_steps),
        )
        self.total_optimizer_updates = (
            self.optimizer_updates_per_epoch * cfg.training.num_epochs
        )
        self.total_scheduler_steps = (
            self.scheduler_steps_per_epoch * cfg.training.num_epochs
        )
        self.scheduler = get_scheduler(
            name=cfg.training.lr_scheduler,
            optimizer=self.optimizer,
            num_warmup_steps=cfg.training.warmup_steps,
            num_training_steps=self.total_scheduler_steps,
        )

        self.accelerator.print(
            "Optimizer updates: "
            f"{self.total_optimizer_updates} | "
            f"Scheduler steps: {self.total_scheduler_steps} "
            f"(epochs={cfg.training.num_epochs}, grad_accum={cfg.training.grad_accum_steps}, "
            f"world_size={self.accelerator.num_processes})"
        )
        if self.logger:
            self.logger.info(f"Optimizer updates: {self.total_optimizer_updates}")
            self.logger.info(f"Scheduler steps: {self.total_scheduler_steps}")
            self.logger.info("Training started.")

        prepared = self.accelerator.prepare(
            self.model,
            self.optimizer,
            self.train_loader,
            self.scheduler,
            *self.dev_loaders.values(),
        )
        self.model = prepared[0]
        self.optimizer = prepared[1]
        self.train_loader = prepared[2]
        self.scheduler = prepared[3]
        dev_loader_names = list(self.dev_loaders.keys())
        self.dev_loaders = dict(zip(dev_loader_names, prepared[4:]))

        self.model.train()
        self.metrics = MetricTracker()
        self.state = TrainingState(
            next_eval_at=cfg.evaluation.interval_samples,
        )

    def _init_tracking(self) -> None:
        if not self.accelerator.is_main_process:
            return

        if self.config_path is not None:
            shutil.copy2(self.config_path, self.output_dir / "train_config.yaml")

        self.logger = TrainLogger(self.output_dir / "train_log.txt")
        self.logger.log_config(self.cfg)

        if self.cfg.logging.wandb_mode == "disabled":
            return

        os.environ["WANDB_MODE"] = self.cfg.logging.wandb_mode
        os.environ["WANDB_DIR"] = str(self.output_dir)
        self.wandb_run = wandb.init(
            project=self.cfg.logging.wandb_project,
            name=self.cfg.output.run_name,
            dir=str(self.output_dir),
            config=self.cfg.to_flat_dict(),
        )

    def _log_wandb(self, metrics: dict[str, float | int], step: int) -> None:
        if self.accelerator.is_main_process and self.wandb_run is not None:
            wandb.log(metrics, step=step)

    def _current_best(self) -> float:
        return self.state.best_mrr if self.state.best_mrr > float("-inf") else 0.0

    def _advance_sample_counters(self) -> None:
        remaining_unique_samples = max(
            len(self.train_dataset) - self.state.epoch_samples_seen,
            0,
        )
        newly_seen = min(self.accelerator.num_processes, remaining_unique_samples)
        self.state.epoch_samples_seen += newly_seen
        self.state.global_samples_seen += newly_seen

    def _run_train_step(
        self,
        batch: dict[str, torch.Tensor],
    ) -> tuple[LossBreakdown, float | None]:
        with self.accelerator.accumulate(self.model):
            micro_bs = self.cfg.training.micro_batch_size
            input_ids = batch["input_ids"]
            attention_mask = batch["attention_mask"]

            logit_parts: list[torch.Tensor] = []
            for chunk_ids, chunk_mask in zip(
                input_ids.split(micro_bs, dim=0),
                attention_mask.split(micro_bs, dim=0),
            ):
                logit_parts.append(
                    extract_yes_no_logits(self.model, chunk_ids, chunk_mask)
                )
            student_z = torch.cat(logit_parts, dim=0).float().unsqueeze(0)
            teacher_scores = (
                batch["teacher_scores"].float().unsqueeze(0).to(student_z.device)
            )
            num_positives: int = batch["num_positives"]
            losses = compute_losses(student_z, teacher_scores, self.cfg, num_positives)

            self.accelerator.backward(losses.total)

            grad_norm: float | None = None
            if self.accelerator.sync_gradients and self.cfg.training.max_grad_norm > 0:
                grad_norm = self.accelerator.clip_grad_norm_(
                    self.model.parameters(),
                    self.cfg.training.max_grad_norm,
                ).item()

            self.optimizer.step()
            self.scheduler.step()
            self.optimizer.zero_grad()

        return losses, grad_norm

    def _emit_train_metrics(
        self,
        *,
        epoch: int,
        grad_norm: float | None,
        progress: tqdm | None,
    ) -> None:
        if self.metrics.count == 0:
            return

        averaged = self.metrics.distributed_averages(self.accelerator)
        current_lr = self.scheduler.get_last_lr()[0]
        payload: dict[str, float | int] = {
            **averaged,
            "lr": current_lr,
            "samples_seen": self.state.global_samples_seen,
            "optimizer_step": self.state.optimizer_steps,
        }
        if grad_norm is not None:
            payload["grad_norm"] = grad_norm

        if self.accelerator.is_main_process:
            self._log_wandb(payload, step=self.state.optimizer_steps)
            if self.logger:
                self.logger.log_step(
                    step=self.state.optimizer_steps,
                    epoch=epoch,
                    samples_seen=self.state.global_samples_seen,
                    learning_rate=current_lr,
                    losses=averaged,
                    grad_norm=grad_norm,
                )
            if progress is not None:
                progress.set_postfix(
                    loss=f"{averaged['loss_total']:.4f}",
                    lr=f"{current_lr:.2e}",
                )

        self.metrics.reset()

    def _maybe_evaluate(self) -> None:
        while self.state.global_samples_seen >= self.state.next_eval_at:
            self.state.next_eval_at += self.cfg.evaluation.interval_samples
            self._run_evaluation(stage="eval")

    def _evaluate_all(self, stage: str) -> tuple[float, dict[str, float]]:
        """Evaluate on all dev datasets, return mean MRR."""
        per_file_mrr: dict[str, float] = {}
        for name, loader in self.dev_loaders.items():
            mrr = evaluate(
                self.model,
                loader,
                self.accelerator,
                micro_batch_size=self.cfg.training.micro_batch_size,
            )
            per_file_mrr[name] = mrr

        mean_mrr = sum(per_file_mrr.values()) / len(per_file_mrr)
        return mean_mrr, per_file_mrr

    def _run_evaluation(self, stage: str) -> float:
        self.accelerator.print(
            f"\n[{stage.capitalize()}] samples_seen={self.state.global_samples_seen}"
        )
        mean_mrr, per_file_mrr = self._evaluate_all(stage)
        is_best = mean_mrr > self.state.best_mrr
        if is_best:
            self.state.best_mrr = mean_mrr

        if self.accelerator.is_main_process:
            for name, mrr in per_file_mrr.items():
                self.accelerator.print(
                    f"  [{stage.capitalize()}] {name}: MRR={mrr:.4f}"
                )
            self.accelerator.print(
                f"[{stage.capitalize()}] mean_MRR={mean_mrr:.4f} "
                f"best={self._current_best():.4f}"
            )
            metric_name = "final_mrr" if stage == "final" else "dev_mrr"
            wandb_payload: dict[str, float | int] = {
                metric_name: mean_mrr,
                "best_mrr": self._current_best(),
                "samples_seen": self.state.global_samples_seen,
            }
            for name, mrr in per_file_mrr.items():
                wandb_payload[f"{metric_name}/{name}"] = mrr
            self._log_wandb(wandb_payload, step=self.state.optimizer_steps)

            if is_best:
                save_model_bundle(
                    self.model,
                    self.tokenizer,
                    self.cfg,
                    self.output_dir / "best",
                    self.accelerator,
                )
            if self.cfg.output.save_every_eval and stage != "final":
                save_model_bundle(
                    self.model,
                    self.tokenizer,
                    self.cfg,
                    self.output_dir / f"samples-{self.state.global_samples_seen}",
                    self.accelerator,
                )
            if self.logger:
                self.logger.log_eval(
                    step=self.state.optimizer_steps,
                    samples_seen=self.state.global_samples_seen,
                    mrr=mean_mrr,
                    best_mrr=self._current_best(),
                    is_best=is_best,
                    stage=stage,
                )
                for name, mrr in per_file_mrr.items():
                    self.logger.info(f"  [{stage.upper()}] {name}: MRR={mrr:.6f}")

            save_eval_results_xlsx(
                self.output_dir / "evaluation_result.xlsx",
                f"samples-{self.state.global_samples_seen}",
                per_file_mrr,
            )

        self.accelerator.wait_for_everyone()
        return mean_mrr

    def train(self) -> None:
        for epoch_index in range(self.cfg.training.num_epochs):
            epoch_number = epoch_index + 1
            self.state.epoch_samples_seen = 0
            progress = tqdm(
                self.train_loader,
                desc=f"Epoch {epoch_number}/{self.cfg.training.num_epochs}",
                disable=not self.accelerator.is_main_process,
            )
            last_grad_norm: float | None = None

            for batch in progress:
                losses, last_grad_norm = self._run_train_step(batch)
                self.metrics.update(losses)
                self._advance_sample_counters()

                if self.accelerator.sync_gradients:
                    self.state.optimizer_steps += 1
                    if (
                        self.state.optimizer_steps % self.cfg.logging.log_interval_steps
                        == 0
                    ):
                        self._emit_train_metrics(
                            epoch=epoch_number,
                            grad_norm=last_grad_norm,
                            progress=progress,
                        )

                self._maybe_evaluate()

            self._emit_train_metrics(
                epoch=epoch_number,
                grad_norm=last_grad_norm,
                progress=progress,
            )

        final_mrr = self._run_evaluation(stage="final")

        if self.accelerator.is_main_process:
            if self.cfg.output.save_last_checkpoint:
                save_model_bundle(
                    self.model,
                    self.tokenizer,
                    self.cfg,
                    self.output_dir / "last",
                    self.accelerator,
                )
            self._log_wandb(
                {"best_mrr": self._current_best()}, step=self.state.optimizer_steps
            )
            if self.wandb_run is not None:
                wandb.finish()
            if self.logger:
                self.logger.info(
                    f"Training complete. Final MRR={final_mrr:.6f} | Best MRR={self._current_best():.6f}"
                )
                self.logger.close()

        self.accelerator.wait_for_everyone()
        self.accelerator.print("Training complete.")
