"""Write evaluation results to a compact xlsx comparison table."""

from __future__ import annotations

from collections.abc import Mapping
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .metrics import mean_score

SHEET_TITLE = "NDCG@10"
AVERAGE_LABEL = "AVERAGE"


def save_results(
    results: Mapping[str, float],
    model_name: str,
    output_path: Path,
) -> None:
    """Write or update one model column in the xlsx summary."""
    table = _load_existing_table(output_path)
    table.setdefault(model_name, {})
    table[model_name].update(
        {dataset: float(score) for dataset, score in results.items()}
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_TITLE

    _write_table(worksheet, table)
    _format_sheet(worksheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"\nResults saved to: {output_path}")


def _load_existing_table(output_path: Path) -> dict[str, dict[str, float]]:
    """Read the current workbook into a model -> dataset -> score mapping."""
    if not output_path.exists():
        return {}

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook.active

    headers = [
        str(worksheet.cell(row=1, column=column).value)
        for column in range(2, worksheet.max_column + 1)
        if worksheet.cell(row=1, column=column).value
    ]

    table: dict[str, dict[str, float]] = {header: {} for header in headers}
    for row in range(2, worksheet.max_row + 1):
        dataset = worksheet.cell(row=row, column=1).value
        if not dataset or str(dataset) == AVERAGE_LABEL:
            continue
        dataset_name = str(dataset)
        for offset, model_name in enumerate(headers, start=2):
            value = worksheet.cell(row=row, column=offset).value
            if isinstance(value, (int, float)):
                table[model_name][dataset_name] = float(value)
    return table


def _write_table(worksheet, table: Mapping[str, Mapping[str, float]]) -> None:
    """Render a complete workbook sheet from a model comparison table."""
    model_names = list(table)
    dataset_names = sorted(
        {
            dataset_name
            for dataset_scores in table.values()
            for dataset_name in dataset_scores
        }
    )

    worksheet.cell(row=1, column=1, value="Dataset")
    for column, model_name in enumerate(model_names, start=2):
        worksheet.cell(row=1, column=column, value=model_name)

    for row, dataset_name in enumerate(dataset_names, start=2):
        worksheet.cell(row=row, column=1, value=dataset_name)
        for column, model_name in enumerate(model_names, start=2):
            score = table[model_name].get(dataset_name)
            if score is not None:
                worksheet.cell(row=row, column=column, value=round(score, 6))

    average_row = len(dataset_names) + 2
    worksheet.cell(row=average_row, column=1, value=AVERAGE_LABEL)
    for column, model_name in enumerate(model_names, start=2):
        values = [
            table[model_name][dataset]
            for dataset in dataset_names
            if dataset in table[model_name]
        ]
        if values:
            worksheet.cell(
                row=average_row, column=column, value=round(mean_score(values), 6)
            )


def _format_sheet(worksheet) -> None:
    """Apply table formatting for readability."""
    header_fill = PatternFill(
        start_color="DCE6F1",
        end_color="DCE6F1",
        fill_type="solid",
    )
    bold = Font(bold=True)

    for column in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=column)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in range(2, worksheet.max_row + 1):
        label = worksheet.cell(row=row, column=1).value
        if label == AVERAGE_LABEL:
            for column in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=column).font = bold
            continue
        worksheet.cell(row=row, column=1).alignment = Alignment(horizontal="left")

    worksheet.column_dimensions["A"].width = 55
    for column in range(2, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(column)].width = 22

    for row in range(2, worksheet.max_row + 1):
        for column in range(2, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=column)
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"
                cell.alignment = Alignment(horizontal="center")
